from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from datetime import datetime, timedelta
import random

wb = Workbook()
wb.remove(wb.active)

# Theme colors
HEADER_COLOR = "1F4E78"
PROFIT_COLOR = "00B050"
LOSS_COLOR = "FF0000"
LIGHT_GRAY = "F2F2F2"

def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws['A1'] = "Trading Performance Dashboard"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", start_color=HEADER_COLOR)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    
    # Report period
    ws['A2'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')
    
    # Key Metrics Section
    ws['A4'] = "Key Performance Metrics"
    ws['A4'].font = Font(size=14, bold=True, color=HEADER_COLOR)
    ws.merge_cells('A4:C4')
    
    metrics = [
        ("Total Return", "='Performance Metrics'!B2", "='Performance Metrics'!C2"),
        ("Annualized Return", "='Performance Metrics'!B3", "='Performance Metrics'!C3"),
        ("Sharpe Ratio", "='Performance Metrics'!B5", ""),
        ("Sortino Ratio", "='Performance Metrics'!B6", ""),
        ("Max Drawdown", "='Performance Metrics'!B9", "='Performance Metrics'!C9"),
        ("Win Rate", "='Performance Metrics'!B12", ""),
        ("Profit Factor", "='Performance Metrics'!B13", ""),
        ("Total Trades", "='Performance Metrics'!B19", ""),
    ]
    
    row = 5
    for metric, value_formula, pct_formula in metrics:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value_formula
        if pct_formula:
            ws[f'C{row}'] = pct_formula
        row += 1
    
    # Format numbers
    for r in range(5, 13):
        ws[f'B{r}'].number_format = '#,##0.00'
        if ws[f'C{r}'].value:
            ws[f'C{r}'].number_format = '0.00%'
    
    # Trade Statistics
    ws['E4'] = "Trade Statistics"
    ws['E4'].font = Font(size=14, bold=True, color=HEADER_COLOR)
    ws.merge_cells('E4:F4')
    
    trade_stats = [
        ("Winning Trades", "='Performance Metrics'!B20"),
        ("Losing Trades", "='Performance Metrics'!B21"),
        ("Current Streak", "='Performance Metrics'!B24"),
        ("Longest Win Streak", "='Performance Metrics'!B25"),
        ("Longest Loss Streak", "='Performance Metrics'!B26"),
        ("Avg Holding Time (hrs)", "='Performance Metrics'!B27"),
    ]
    
    row = 5
    for stat, formula in trade_stats:
        ws[f'E{row}'] = stat
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'F{row}'] = formula
        ws[f'F{row}'].number_format = '#,##0.00'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

def create_trade_log(wb):
    ws = wb.create_sheet("Trade Log")
    
    headers = ["Trade ID", "Symbol", "Direction", "Entry Date", "Exit Date", 
               "Entry Price", "Exit Price", "Quantity", "PnL", "PnL %", "Holding Time (hrs)", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data
    symbols = ["BTC/USDT", "ETH/USDT", "SOL/USDT", "BNB/USDT", "ADA/USDT", "AVAX/USDT"]
    directions = ["LONG", "SHORT"]
    
    start_date = datetime.now() - timedelta(days=90)
    
    for i in range(2, 22):
        symbol = random.choice(symbols)
        direction = random.choice(directions)
        entry_date = start_date + timedelta(days=random.randint(0, 85))
        holding_hours = random.randint(2, 72)
        exit_date = entry_date + timedelta(hours=holding_hours)
        
        entry_price = random.uniform(100, 50000)
        price_change = random.uniform(-0.15, 0.25)
        exit_price = entry_price * (1 + price_change)
        quantity = random.uniform(0.1, 10)
        
        ws[f'A{i}'] = f"TRD{i-1:04d}"
        ws[f'B{i}'] = symbol
        ws[f'C{i}'] = direction
        ws[f'D{i}'] = entry_date
        ws[f'E{i}'] = exit_date
        ws[f'F{i}'] = entry_price
        ws[f'G{i}'] = exit_price
        ws[f'H{i}'] = quantity
        ws[f'I{i}'] = f"=(G{i}-F{i})*H{i}" if direction == "LONG" else f"=(F{i}-G{i})*H{i}"
        ws[f'J{i}'] = f"=(G{i}-F{i})/F{i}" if direction == "LONG" else f"=(F{i}-G{i})/F{i}"
        ws[f'K{i}'] = holding_hours
        ws[f'L{i}'] = "CLOSED"
    
    # Formatting
    for row in range(2, 22):
        ws[f'D{row}'].number_format = 'yyyy-mm-dd hh:mm'
        ws[f'E{row}'].number_format = 'yyyy-mm-dd hh:mm'
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'H{row}'].number_format = '0.0000'
        ws[f'I{row}'].number_format = '#,##0.00'
        ws[f'J{row}'].number_format = '0.00%'
        ws[f'K{row}'].number_format = '0.00'
    
    # Column widths
    widths = [12, 12, 10, 18, 18, 12, 12, 10, 12, 10, 18, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    
    # Totals row
    ws['H22'] = "TOTALS:"
    ws['H22'].font = Font(bold=True)
    ws['I22'] = "=SUM(I2:I21)"
    ws['I22'].font = Font(bold=True)
    ws['I22'].number_format = '#,##0.00'

def create_performance_metrics(wb):
    ws = wb.create_sheet("Performance Metrics")
    
    ws['A1'] = "Performance Metrics Summary"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", start_color=HEADER_COLOR)
    ws.merge_cells('A1:C1')
    
    metrics_data = [
        ("Returns", ""),
        ("Total Return", "=SUM('Trade Log'!I2:I21)", "=B2/10000"),
        ("Annualized Return", "=B2/(('Trade Log'!E21-'Trade Log'!D2)/365)", "=B3/10000"),
        ("", ""),
        ("Risk-Adjusted Returns", ""),
        ("Sharpe Ratio", 1.85, ""),
        ("Sortino Ratio", 2.34, ""),
        ("Calmar Ratio", 1.52, ""),
        ("", ""),
        ("Drawdown", ""),
        ("Max Drawdown", -1250.50, "-12.51%"),
        ("Current Drawdown", -320.00, "-3.20%"),
        ("", ""),
        ("Win/Loss Statistics", ""),
        ("Win Rate", "=COUNTIFS('Trade Log'!I2:I21,\">0\")/COUNTA('Trade Log'!I2:I21)", ""),
        ("Profit Factor", "=SUMIF('Trade Log'!I2:I21,\">0\")/ABS(SUMIF('Trade Log'!I2:I21,\"<0\"))", ""),
        ("Average Win", "=AVERAGEIF('Trade Log'!I2:I21,\">0\")", ""),
        ("Average Loss", "=AVERAGEIF('Trade Log'!I2:I21,\"<0\")", ""),
        ("Largest Win", "=MAX('Trade Log'!I2:I21)", ""),
        ("Largest Loss", "=MIN('Trade Log'!I2:I21)", ""),
        ("", ""),
        ("Trade Statistics", ""),
        ("Total Trades", "=COUNTA('Trade Log'!A2:A21)", ""),
        ("Winning Trades", "=COUNTIF('Trade Log'!I2:I21,\">0\")", ""),
        ("Losing Trades", "=COUNTIF('Trade Log'!I2:I21,\"<0\")", ""),
        ("Expectancy", "=AVERAGE('Trade Log'!I2:I21)", ""),
        ("Expectancy %", "=AVERAGE('Trade Log'!J2:J21)", ""),
        ("Current Streak", 3, ""),
        ("Longest Win Streak", 5, ""),
        ("Longest Loss Streak", 3, ""),
        ("Avg Holding Time (hrs)", "=AVERAGE('Trade Log'!K2:K21)", ""),
    ]
    
    row = 2
    for label, value, pct in metrics_data:
        ws[f'A{row}'] = label
        if label and not value:
            ws[f'A{row}'].font = Font(bold=True, size=12, color=HEADER_COLOR)
        else:
            ws[f'B{row}'] = value
            ws[f'C{row}'] = pct
        row += 1
    
    # Formatting
    for r in range(2, row):
        if ws[f'B{r}'].value and isinstance(ws[f'B{r}'].value, str) and ws[f'B{r}'].value.startswith('='):
            ws[f'B{r}'].number_format = '#,##0.00'
        if ws[f'C{r}'].value and isinstance(ws[f'C{r}'].value, str) and (ws[f'C{r}'].value.startswith('=') or '%' in str(ws[f'C{r}'].value)):
            ws[f'C{r}'].number_format = '0.00%'
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15

def create_monthly_analysis(wb):
    ws = wb.create_sheet("Monthly Analysis")
    
    headers = ["Month", "Trades", "Win Rate", "Total PnL", "Return %", "Best Trade", "Worst Trade"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
    
    months = ["2026-01", "2026-02", "2026-03"]
    
    for i, month in enumerate(months, 2):
        ws[f'A{i}'] = month
        ws[f'B{i}'] = random.randint(5, 10)
        ws[f'C{i}'] = random.uniform(0.5, 0.8)
        ws[f'D{i}'] = random.uniform(-500, 2000)
        ws[f'E{i}'] = random.uniform(-0.05, 0.20)
        ws[f'F{i}'] = random.uniform(200, 800)
        ws[f'G{i}'] = random.uniform(-400, -50)
    
    for row in range(2, 5):
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '0.00%'
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].number_format = '#,##0.00'
    
    for col in range(1, 8):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 15

def create_risk_analysis(wb):
    ws = wb.create_sheet("Risk Analysis")
    
    ws['A1'] = "Risk Analysis 
